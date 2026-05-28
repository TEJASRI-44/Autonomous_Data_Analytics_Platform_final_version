import pandas as pd


import pandas as pd
import csv
from openpyxl import load_workbook


def load_dataset(file_path):

    try:


        if file_path.endswith(".csv"):

            try:


                with open(

                    file_path,
                    "r",
                    encoding="utf-8"

                ) as file:

                    sample = file.read(2048)

                    detected_dialect = (

                        csv.Sniffer()
                        .sniff(sample)
                    )

                    delimiter = (
                        detected_dialect.delimiter
                    )


                df = pd.read_csv(

                    file_path,
                    sep=delimiter
                )

            except Exception:

                raise ValueError(

                    "CSV content could not "
                    "be converted into "
                    "structured tabular data"
                )


        elif file_path.endswith(".xlsx"):

            try:


                workbook = load_workbook(

                    file_path,
                    data_only=True
                )

                active_sheet = workbook.active

                # Check if sheet has data

                if active_sheet.max_row < 2:

                    raise ValueError(
                        "Excel sheet is empty"
                    )

                df = pd.read_excel(file_path)

            except Exception:

                raise ValueError(

                    "Excel content could not "
                    "be parsed into "
                    "structured tabular data"
                )

        else:

            raise ValueError(
                "Unsupported file format"
            )

        # Remove fully empty rows

        df = df.dropna(how="all")

        # Remove fully empty columns

        df = df.dropna(
            axis=1,
            how="all"
        )

        # Normalize column names

        df.columns = [

            str(col).strip()

            for col in df.columns
        ]

        if df.empty:

            raise ValueError(
                "Dataset is empty"
            )

        if len(df.columns) < 2:

            raise ValueError(

                "Dataset appears "
                "unstructured or malformed"
            )

        return df

    except Exception as e:

        raise ValueError(

            f"Invalid dataset content: {str(e)}"
        )

